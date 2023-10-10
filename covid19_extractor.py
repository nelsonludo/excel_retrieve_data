import math
import pandas as pd
import locale
from datetime import date, timedelta
import xlsxwriter
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os


# Set the locale to French
locale.setlocale(locale.LC_TIME, 'fr_FR')


# Load the Excel file
input_file_path = 'X:\\Unite TIC\\Nelson\\Synthese_rapports_regions.xlsx'
df = pd.read_excel(input_file_path)

# Get the wednesday of current week
# Get the current date
current_date = date.today()

# Get the week number
week_number = current_date.strftime("%W")

# Get the date of the Wednesday in the current week
days_ahead = (2 - current_date.weekday()) % 7
wednesday_date = current_date + timedelta(days=days_ahead)

#target_value = 'Rapports des régions du 27 Septembre 2023'
target_date = pd.to_datetime('2023-10-04')

# Initialize variables to store the row and column indices of the target cell
target_row_index = -1
target_column_index = -1

# Iterate through each cell in the DataFrame
for row_index, row in df.iterrows():
    for column_index, value in row.items():
        # Check if the cell value is a string and if it contains the target date
        if pd.isna(value):
            continue
        elif isinstance(value, str) :
            value_lower = value.lower() #convert the cell value to lower case 
            formatted_month = target_date.strftime('%B').lower() #convert the target date to lower case
            target_date_formatted = target_date.strftime('%d ') + formatted_month + target_date.strftime(' %Y')
            
            #compare the cell value and the target date
            if target_date_formatted in value_lower:
                target_row_index = row_index
                target_column_index = column_index
            break
    if target_row_index != -1:
        break

# Reset the locale back to the default
locale.setlocale(locale.LC_TIME, '')


#print(target_date.strftime('%d %B %Y'))

# Check if the target cell was found
if target_row_index == -1 and target_column_index == -1:
    print("No cell with the target date found.")
    # Calculate the range of cells to retrieve based on the target cell
else:
    print(f"Target cell found at row {target_row_index} and column {target_column_index}.")

    #print("target_row:", target_row_index)
    #print("target_column:", target_column_index)

    #Set the number of columns and rows to retrieve after the target cell
    num_columns_to_retrieve = 12
    num_rows_to_retrieve = 13



    # Get the label of the row containing the target cell
    start_row = df.index[target_row_index] 
    #print("start_row:", start_row)
    end_row = start_row + num_rows_to_retrieve
    start_column = target_column_index

    start_column_index = df.columns.get_loc(start_column)
    end_column_index = start_column_index + num_columns_to_retrieve
    end_column = end_column_index

    
    # Get the index of the column containing the second target cell
    second_start_column = start_column_index + 15
    second_end_column= second_start_column + num_columns_to_retrieve

    # Get the label of the row containing the target cell of the previous week 
    start_row_last_week = df.index[row_index - 15   ]
    #print("start_row_last_week:", start_row_last_week)
    end_row_last_week = start_row_last_week + num_rows_to_retrieve + 2

    #print(type(start_column_index))
    #print(end_column)

    # Extract the desired range of cells

    extracted_data = df.iloc[start_row:end_row, start_column_index:end_column]
    extracted_data2 = df.iloc[start_row:end_row, second_start_column:second_end_column]
    extracted_data3 = df.iloc[start_row_last_week:end_row_last_week, second_start_column:second_end_column]

    # Create a new Excel file
    output_file_path = 'X:\\Unite TIC\\Nelson\\synthese_covid.xlsx'
    writer = pd.ExcelWriter(output_file_path)

    extracted_data.columns = pd.Index([None] * len(extracted_data.columns)) 
    extracted_data.columns = extracted_data.iloc[1]   
    
    extracted_data2.columns = pd.Index([None] * len(extracted_data2.columns)) 
    extracted_data2.columns = extracted_data2.iloc[1]   
    
    extracted_data3.columns = pd.Index([None] * len(extracted_data3.columns)) 
    extracted_data3.columns = extracted_data3.iloc[1]   

    # Write the extracted data to the new file
    extracted_data.to_excel(writer, sheet_name='this week', index=False, columns=None)
    extracted_data2.to_excel(writer, sheet_name='this week', index=False, startcol=extracted_data.shape[1]+2, columns=None)
    extracted_data3.to_excel(writer, sheet_name='this week', index=False, startrow=extracted_data.shape[0]+3, columns=None)

    # Create a new Excel file
    workbook = writer.book  #Workbook()
    worksheet = writer.sheets['this week'] #workbook.active

    # Define the table range and table format for each extracted data
    

     # Load the Excel file
    workbook1 = openpyxl.load_workbook(input_file_path)

    # Get the active sheet (you can also specify a specific sheet by name)
    sheet = workbook1.active
    table_name = "table1"
    # Iterate over all tables in the sheet
    for table in sheet.tables.values():
        table_ranged = table.ref  # Get the range of the table (e.g., 'A1:C10')
        table_named = table.name  # Get the name of the table
        table_cellsd = sheet[table_ranged]  # Get the cells within the table range

        # Convert column index to letter
        target_column_letter = get_column_letter(start_column_index+1)
        target_row_for_table_name = target_row_index+3

        # Check if the extracted cells belong to the table
        for row in table_cellsd:
            for cell in row:
                if cell.coordinate == f'{target_column_letter}{target_row_for_table_name}':  # Example: check if cell A1 belongs to the table
                    print(f"The cell {cell.coordinate} belongs to the table '{table_named}'")
                    table_name = table_named
                else :
                    print("the cell does not belong to any table known to man")
    #print(f'{target_column_letter}{target_row_for_table_name}')
    # Close the workbook
    workbook1.close()
 
    table_range = f'A3:{chr(65 + extracted_data.shape[1] - 1)}{extracted_data.shape[0]+1}'
    table_format = workbook.add_format({'border': 1, 'align': 'center'})
    
    
    # Create a list of column names from the extracted_data DataFrame starting from the specified row
    column_names = extracted_data.columns.tolist()

    # Create a list of dictionaries for each column in the table
    columns_data = [{'header': name} for name in column_names]
    
    #trying to find out why the code doesn't work
    worksheet.add_table(table_range, {'name': table_name, 'style': 'Table Style Medium 9', 'columns': columns_data})



    workbook2 = openpyxl.load_workbook(input_file_path)

    # Get the active sheet (you can also specify a specific sheet by name)
    sheet = workbook2.active

    table_name2 = "table2"
    # Iterate over all tables in the sheet
    for table in sheet.tables.values():
        table_ranged = table.ref  # Get the range of the table (e.g., 'A1:C10')
        table_named = table.name  # Get the name of the table
        table_cellsd = sheet[table_ranged]  # Get the cells within the table range

        # Convert column index to letter
        target_column_letter = get_column_letter(second_start_column+1)
        target_row_for_table_name = target_row_index+3

        # Check if the extracted cells belong to the table
        for row in table_cellsd:
            for cell in row:
                if cell.coordinate == f'{target_column_letter}{target_row_for_table_name}':  # Example: check if cell A1 belongs to the table
                    print(f"The cell {cell.coordinate} belongs to the table '{table_named}'")
                    table_name2 = table_named
                else :
                    print("the cell does not belong to any table known to man")
    #print(f'{target_column_letter}{target_row_for_table_name}')
    # Close the workbook
    workbook2.close()


    table_range2 = f'{xlsxwriter.utility.xl_col_to_name(extracted_data.shape[1] + 2)}3:{xlsxwriter.utility.xl_col_to_name(extracted_data.shape[1] + 1 + extracted_data2.shape[1])}{extracted_data2.shape[0]+1}'
    table_format2 = workbook.add_format({'border': 1, 'align': 'center'})

    
    # Create a list of column names from the extracted_data DataFrame starting from the specified row
    column_names2 = extracted_data2.columns.tolist()

    # Convert all elements to string and replace "nan" with "nothing"
    column_names2_without_nan = [str(name) if str(name) != "nan" else "nothing" for name in column_names2]

    # Create a list of dictionaries for each column in the table
    columns_data2 = [{'header': name} for name in column_names2_without_nan]
    
    worksheet.add_table(table_range2, {'name': table_name2, 'style': 'Table Style Medium 9', 'columns': columns_data2})
    
    
    table_name3 = "table3"
    # Load the Excel file
    workbook3 = openpyxl.load_workbook(input_file_path)

    # Get the active sheet (you can also specify a specific sheet by name)
    sheet = workbook3.active

    # Iterate over all tables in the sheet
    for table in sheet.tables.values():
        table_ranged = table.ref  # Get the range of the table (e.g., 'A1:C10')
        table_named = table.name  # Get the name of the table
        table_cellsd = sheet[table_ranged]  # Get the cells within the table range

        # Convert column index to letter
        target_column_letter = get_column_letter(second_start_column+2)
        target_row_for_table_name = target_row_index-12

        # Check if the extracted cells belong to the table
        for row in table_cellsd:
            for cell in row:
                if cell.coordinate == f'{target_column_letter}{target_row_for_table_name}':  # Example: check if cell A1 belongs to the table
                    print(f"The cell {cell.coordinate} belongs to the table '{table_named}'")
                    table_name3 = table_named
                else :
                    print("the cell does not belong to any table known to man")
    # Close the workbook
    workbook3.close()

    table_range3 = f'A{extracted_data.shape[0] + 6}:{chr(65 + extracted_data.shape[1] - 1)}{extracted_data2.shape[0] + extracted_data3.shape[0] + 2}'
    table_format3 = workbook.add_format({'border': 1, 'align': 'center'})

     # Create a list of column names from the extracted_data DataFrame starting from the specified row
    column_names3 = extracted_data3.columns.tolist()

    # Convert all elements to string and replace "nan" with "nothing"
    column_names3_without_nan = [str(name) if str(name) != "nan" else "nothing" for name in column_names3]

    
    # Create a list of dictionaries for each column in the table
    columns_data3 = [{'header': name} for name in column_names3_without_nan]
    
    

    worksheet.add_table(table_range3, {'name': table_name3, 'style': 'Table Style Medium 9', 'columns': columns_data3}) #, 'columns': columns_data3


    #print(table_range)
    #worksheet.title = 'this week'


    ## Write the extracted data to the worksheet
    #for row in dataframe_to_rows(extracted_data, index=False, header=True):
    #    worksheet.append(row)

     #for row in dataframe_to_rows(extracted_data2, index=False, header=True):
    #    worksheet.append(row)

    #for row in dataframe_to_rows(extracted_data3, index=False, header=True):
    #    worksheet.append(row)

    ## Define the table range for each extracted data
    #table_range = f'A1:{chr(65 + extracted_data.shape[1] - 1)}{extracted_data.shape[0] + 1}'
    #table_range2 = f'{chr(65 + extracted_data.shape[1] + 2)}1:{chr(65 + extracted_data.shape[1] + extracted_data2.shape[1] + 1)}{extracted_data2.shape[0] + 1}'
    #table_range3 = f'{chr(65 + extracted_data.shape[1] + 2)}{extracted_data.shape[0] + 3}:{chr(65 + extracted_data.shape[1] + extracted_data2.shape[1] + 1)}{extracted_data2.shape[0] + extracted_data3.shape[0] + 2}'

    ## Create tables for each extracted data range
    #table1 = Table(displayName="Table1", ref=table_range)
    #table2 = Table(displayName="Table2", ref=table_range2)
    #table3 = Table(displayName="Table3", ref=table_range3)

    ## Set the table style
    #table_style = TableStyleInfo(name="TableStyleMedium9")

    #table1.tableStyleInfo = table_style
    #table2.tableStyleInfo = table_style
    #table3.tableStyleInfo = table_style

    ## Add tables to the worksheet
    #worksheet.add_table(table1)
    #worksheet.add_table(table2)
    #worksheet.add_table(table3)

    ## Save the workbook
    #output_file_path = 'synthese_covid.xlsx'
    #workbook.save(output_file_path)

    print("Data extracted and saved to Excel file successfully.")


    # Save and close the new file
    writer._save()
    writer.close()

    #print(extracted_data3)
    #print(extracted_data.shape)
